
import timeit
from openpyxl import Workbook

def benchmark_get_cell():
    wb = Workbook()
    ws = wb.active
    # Warm up / create some cells
    for r in range(1, 101):
        for c in range(1, 101):
            ws.cell(row=r, column=c)

    def access_cells():
        for r in range(1, 101):
            for c in range(1, 101):
                ws.cell(row=r, column=c)

    timer = timeit.Timer(access_cells)
    iterations = 100
    total_time = timer.timeit(number=iterations)
    print(f"Time to access 10,000 existing cells {iterations} times: {total_time:.4f}s")
    print(f"Average time per access: {total_time / (iterations * 10000) * 1e6:.4f}us")

if __name__ == "__main__":
    benchmark_get_cell()
